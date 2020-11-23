import csv
import json
import re

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from six import string_types

import frappe
from frappe.core.doctype.data_import.importer import Importer, ImportFile
from frappe.model.document import Document
from frappe import _
from frappe.utils import flt
from frappe.utils.background_jobs import enqueue
from frappe.core.page.background_jobs.background_jobs import get_info
from frappe.utils.scheduler import is_scheduler_inactive
from frappe.utils.xlsxutils import handle_html, ILLEGAL_CHARACTERS_RE
from erpnext.accounts.utils import get_balance_on
from erpnext.accounts.report.bank_reconciliation_statement.bank_reconciliation_statement import get_entries, get_amounts_not_reflected_in_system


@frappe.whitelist()
def get_bank_transactions(bank_account, from_date = None, to_date = None):
	filters = []
	filters.append(['bank_account' ,'=', bank_account])
	filters.append(['docstatus' ,'=', 1])
	filters.append(['unallocated_amount' ,'>', 0])
	if to_date:
		filters.append(['date' ,'<=', to_date])
	if from_date:
		filters.append(['date' ,'>=', from_date])
	transactions = frappe.get_list(
		'Bank Transaction',
		fields = ['date', 'debit', 'credit', 'currency',
		'description', 'name', 'bank_account', 'company',
		'reference_number', 'transaction_id'],
		filters = filters
	)
	print(len(transactions))
	return transactions

@frappe.whitelist()
def get_account_balance(bank_account, till_date):
	account = frappe.db.get_value('Bank Account', bank_account, 'account')
	print(account)
	# balance_as_per_system = get_balance_on(account, till_date)
	# print(balance_as_per_system)
	# return balance_as_per_system
	filters = frappe._dict({
		"account": account,
		"report_date": till_date,
		"include_pos_transactions": 1
	})
	print(type(filters))
	data = get_entries(filters)

	balance_as_per_system = get_balance_on(filters["account"], filters["report_date"])

	total_debit, total_credit = 0,0
	for d in data:
		total_debit += flt(d.debit)
		total_credit += flt(d.credit)

	amounts_not_reflected_in_system = get_amounts_not_reflected_in_system(filters)

	bank_bal = flt(balance_as_per_system) - flt(total_debit) + flt(total_credit) \
		+ amounts_not_reflected_in_system

	return bank_bal

@frappe.whitelist()
def get_importer_preview(import_file_path, data_import=None, template_options=None):
	data_import = frappe.get_doc(doctype="Data Import")
	data_import.import_type = "Insert New Records"
	data_import.reference_doctype = "Bank Transaction"
	data_import.import_file = import_file_path
	data_import.submit_after_import = 1
	data_import.template_options = template_options
	importer = Importer("Bank Transaction", data_import=data_import, file_path = import_file_path)
	preview = importer.get_data_for_import_preview()
	return {"preview":preview}


@frappe.whitelist()
def form_start_import(import_file_path, data_import=None, template_options=None, bank_account=None):

	if is_scheduler_inactive() and not frappe.flags.in_test:
		frappe.throw(
			_("Scheduler is inactive. Cannot import data."), title=_("Scheduler Inactive")
		)

	enqueued_jobs = [d.get("job_name") for d in get_info()]
	if "bank_statement_import" not in enqueued_jobs:
		enqueue(
			start_import,
			queue="default",
			timeout=6000,
			event="data_import",
			job_name="bank_statement_import",
			import_file_path=import_file_path,
			data_import=data_import,
			template_options=template_options,
			bank_account=bank_account,
			now=frappe.conf.developer_mode or frappe.flags.in_test,
		)
		return True
	return False


def start_import(import_file_path, template_options=None, bank_account=None):
	"""This method runs in background job"""
	data_import = frappe.get_doc(doctype="Data Import")
	data_import.import_type = "Insert New Records"
	data_import.reference_doctype = "Bank Transaction"
	data_import.import_file = import_file_path
	data_import.submit_after_import = 1
	data_import.template_options = template_options

	import_file = ImportFile("Bank Transaction", file = import_file_path, import_type="Insert New Records")
	data = import_file.raw_data

	bank_account_loc = None

	if "Bank Account" not in data[0]:
		data[0].append("Bank Account")
	else:
		for loc, header in enumerate(data[0]):
			if header == "Bank Account":
				bank_account_loc = loc

	for row in data[1:]:
		if bank_account_loc:
			row[bank_account_loc] = bank_account
		else:
			row.append(bank_account)

	full_file_path = import_file.file_doc.get_full_path()
	parts = import_file.file_doc.get_extension()
	extension = parts[1]
	extension = extension.lstrip(".")

	if extension == "csv":
		with open(full_file_path, 'w', newline='') as file:
			writer = csv.writer(file)
			writer.writerows(data)
	elif extension == "xlsx" or "xls":
		write_xlsx(data, "trans", file_path = full_file_path)

	try:
		i = Importer("Bank Transaction", data_import=data_import, file_path = import_file_path)
		i.import_data()
	except Exception:
		frappe.db.rollback()
		data_import.db_set("status", "Error")
		frappe.log_error(title="data_import")
	finally:
		frappe.flags.in_import = False

def write_xlsx(data, sheet_name, wb=None, column_widths=None, file_path=None):
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook(write_only=True)

	ws = wb.create_sheet(sheet_name, 0)

	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	row1 = ws.row_dimensions[1]
	row1.font = Font(name='Calibri', bold=True)

	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, string_types) and (sheet_name not in ['Data Import Template', 'Data Export']):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, string_types) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = re.sub(ILLEGAL_CHARACTERS_RE, '', value)

			clean_row.append(value)

		ws.append(clean_row)

	wb.save(file_path)
	return True

@frappe.whitelist()
def update_bank_transaction(bank_transaction, transaction_id, reference_number, party_type, party):

	updates = {
		"transaction_id": transaction_id,
		"reference_number": reference_number,
		"party_type": party_type,
		"party": party,
	}
	frappe.db.set_value("Bank Transaction", bank_transaction, updates)

@frappe.whitelist()
def create_payment_entry_bts(
		bank_transaction,
		transaction_id,
		reference_number,
		reference_date,
		party_type,
		party,
		posting_date,
		mode_of_payment,
		project,
		cost_center
	):
	bank_transaction = frappe.get_doc("Bank Transaction", bank_transaction)
	paid_amount = bank_transaction.credit if bank_transaction.credit > 0 else bank_transaction.debit
	payment_type = "Receive" if bank_transaction.credit > 0 else "Pay"
	party_type = "Customer" if bank_transaction.credit > 0 else "Supplier"

	company_account = frappe.get_value("Bank Account", bank_transaction.bank_account, "account")
	payment_entry = frappe.new_doc("Payment Entry")
	payment_entry.company = "Moha"
	payment_entry.payment_type = payment_type
	payment_entry.transaction_id =  transaction_id
	payment_entry.reference_no =  reference_number
	payment_entry.reference_date =  reference_date
	payment_entry.party_type =  party_type
	payment_entry.party =  party
	payment_entry.posting_date =  posting_date
	payment_entry.mode_of_payment =  mode_of_payment
	payment_entry.project =  project
	payment_entry.cost_center =  cost_center
	payment_entry.paid_amount = paid_amount
	payment_entry.received_amount = paid_amount

	if payment_type == "Receive":
		payment_entry.paid_to = company_account
	else:
		payment_entry.paid_from = company_account
	payment_entry.insert()
	payment_entry.submit()

	return payment_entry.name
